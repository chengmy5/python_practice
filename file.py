# f=open('E:/我司具有资质.txt', 'r', encoding='utf-8')
# content=f.read()
# qualification=content.split(',')
# print(qualification)
# f.close()
import os
# a=os.path.exists('E:file')
# b=os.path.exists('E:/bidding_information')
# os.makedirs('E:/bidding_information')
# print(a)
# print(b)

# from shutil import copy
#
# copy('E:/file/aa.txt','E:/file2')



# import requests
# import bs4
# from requests.packages.urllib3.exceptions import InsecureRequestWarning
#
# requests.packages.urllib3.disable_warnings(InsecureRequestWarning)
#
# url='https://b2b.10086.cn/b2b/main/showBiao!showKaibiaoResult.html?ekp1APd1=5r4ygCLbjMRcG2jrQVDc_Z9gsfEszfyhsVf1ytXD15e5TsZfRs6V9UiFU1L8Xf79f55NxUyCaMd4TdebWNuGuQ1IL9bBkhWd48H87BrdBxKKuQ3LAnRU2uLFn0_1FzZaQ_Q2Z5X0oMiKGLPNH2hnNKdqThhqnIHlCDlkJUwNLKk.PvWpdgOTvLFNMkj593beUoRfoY8KuVrmOlbe5aLPsz5AP44VUZM92hlUXoWNCXIb981sp3WmRWQ7vjvKtnVuBJ2TzyFZnKwQ3ODxDknxE5vr0rvC2yibLl6aOG5HviZLjjb4oryZYcsn5.VUPDixJke7ySRjnjCbJ4lB8TDhnJdlJLUNrNK01Ea7IsjZ2cm7'
# data={}
# data['page.currentPage']='1'
# data['page.perPageSize']='20'
# data['noticeBean.companyName']=''
# data['noticeBean.title']=''
# data['noticeBean.startDate']='2020-12-01'
# data['noticeBean.endDate']='2021-01-02'
# res = requests.post(url,data,verify=False,headers={'User-Agent':'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36 Edg/87.0.664.66',
#                                         'Cookie':'e01dcfc741ea014fb8=c08f0bdae7544ade9c4a3eb2a1f65840; saplb_*=(J2EE204290120)204290150; JSESSIONID=C4JaobR7R6oMSaktqddqIctcfS3BdgFmOC0M_SAP1TTcLgkNGtkASkg-T5gPsSgf; LS2cj8YxOIT5O=5zxbClB9UMBUvLNXJpy3sThzSUo6Kfp7SxTqPuhDUlH_mZDPUk9kxALIrU5JUfKqI2WZW_XDwLtMZ.IKGoFAPeA; LS2cj8YxOIT5P=5U5dv3mdB_aWqqqmCa1cW8q2ZN5FN9HzvDTOS2TlbL7IAcMCqU2XMbDntmD6T651jmRU7cw8elTBR2czBxbzYyI9mh18aTYIUedbc2oBs7uaW2vyIwj1gGPT3FK04zYP_e0rERhiJFTCt_ldtokkCTTiPev4NSz8xS8tVcaPpSZViS0WB.SviWLEbZW9dubiOG15Un5hXT_KyDtRiO_klOk2l.BwwYCACC147iaAGwHrry3.NRJV5lCktRIJouVCb8bfFujn8bU4b.f40eun78FcsetZLRXnm_WtfEBt.i.Y0N2L.QjFkZER7MTMqVSemW'})
# soup = bs4.BeautifulSoup(res.text, 'html.parser')
# print(res)

a=[{
    "featureType": "background",
    "elementType": "geometry",
    "stylers": {
        "visibility": "on"
    }
}, {
    "featureType": "manmade",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "manmade",
    "elementType": "geometry",
    "stylers": {
        "visibility": "on"
    }
}, {
    "featureType": "education",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "medical",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "scenicspots",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "scenicspots",
    "elementType": "geometry",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "medical",
    "elementType": "geometry",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "education",
    "elementType": "geometry",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "subwaystation",
    "elementType": "geometry",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "entertainment",
    "elementType": "geometry",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "estate",
    "elementType": "geometry",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "shopping",
    "elementType": "geometry",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "transportation",
    "elementType": "geometry",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "transportation",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "playground",
    "elementType": "geometry",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "parkinglot",
    "elementType": "geometry",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "road",
    "elementType": "labels",
    "stylers": {
        "visibility": "on"
    }
}, {
    "featureType": "highway",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "nationalway",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "provincialway",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "cityhighway",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "arterial",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "fourlevelway",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "local",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "universityway",
    "elementType": "geometry",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "subway",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "subway",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "highwaysign",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "subwaylabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "parkingspace",
    "elementType": "geometry",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "airportlabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "airportlabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "scenicspotslabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "scenicspotslabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "educationlabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "medicallabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "medicallabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "entertainmentlabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "entertainmentlabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "estatelabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "estatelabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "businesstowerlabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "businesstowerlabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "companylabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "companylabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "governmentlabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "on"
    }
}, {
    "featureType": "governmentlabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "on"
    }
}, {
    "featureType": "restaurantlabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "restaurantlabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "hotellabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "hotellabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "shoppinglabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "shoppinglabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "lifeservicelabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "lifeservicelabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "carservicelabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "carservicelabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "transportationlabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "transportationlabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "financelabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "financelabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "districtlabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "on"
    }
}, {
    "featureType": "districtlabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "on"
    }
}, {
    "featureType": "poilabel",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "poilabel",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "road",
    "elementType": "geometry",
    "stylers": {
        "visibility": "on"
    }
}, {
    "featureType": "village",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "town",
    "elementType": "labels.icon",
    "stylers": {
        "visibility": "off"
    }
}, {
    "featureType": "town",
    "elementType": "labels",
    "stylers": {
        "visibility": "off"
    }
}]
for each in a:
    print(each)

import openpyxl
import random

# print(str(random.uniform(0.32,0.3501))[:5])

# wb=openpyxl.load_workbook(r'E:\测试资料站点清单\ZTSJ_25测试资料.xlsx')
# ws=wb['Sheet1']
# name=ws['A']
# le=ws['B']
#
# a=[]
# b=[]
# for n in name:
#     a.append(n.value)
# for n in le:
#     b.append(n.value)
# a.pop(0)
# b.pop(0)
# print(a)
# print(b)


# wa=sheet_name['A2':'A10']
# for i in wa:
#     for j in i:
#
#         print(j.value)



import os
# print(os.listdir('E:\测试资料站点清单'))
# import requests
# import bs4
#
#
# url="https://b2b.10086.cn/b2b/main/showBiao!showZhaobiaoResult.html?page.currentPage=1&page.perPageSize=20&noticeBean.companyName=&noticeBean.title=&noticeBean.startDate=&noticeBean.endDate="
# data = {}
# data['page.currentPage']='1'
# data['page.perPageSize']='20'
# data['noticeBean.companyName']=''
# data['noticeBean.title']=''
# data['noticeBean.startDate']=''
# data['noticeBean.endDate']=''
# headers={'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
#          'Referer': 'https://b2b.10086.cn/b2b/main/showBiao!preShowBiao.html?noticeType=list1'}
# res = requests.post(url,data,verify=False,headers=headers)
# soup = bs4.BeautifulSoup(res.text, 'html.parser')
# print(soup)
# import re
#
# a='	沙角A电厂2021年4、5号机组B、C级检修锅炉脱硝系统SCR第一层初装层催化剂更换工作项目（第二次招标）招标公告'
# regex=re.search("[^\u4e00-\u9fa5]",a)
# print('\t')
# print(a.replace(regex.group(),''))

# from selenium import webdriver
#
# driver_path=r'D:\driver\IEDriverServer.exe'
# driver=webdriver.Ie(executable_path=driver_path)
# url="http://baidu.com"
#
# driver.get(url)
# content=driver.page_source
# print(content)


# import time
# print(time.strftime('%Y%m%d',time.localtime()))
# wb=openpyxl.Workbook()
# ws=wb.active
# ws.append(['序号','招标网站','项目名称','招标单位','项目概况','报名截止时间','资质要求','招标范围'])
# wb.save('E:\每日招标信息01.xlsx')
# wb=openpyxl.open('E:\每日招标信息01.xlsx')
# ws=wb.active
# i=0
# while i<10:
#     ws.append(['1','工信部','xxx招标项目','xxx科技有限公司','400万','2021-02-05','xxx一级','广东'])
#     i+=1
# wb.save('E:\每日招标信息01.xlsx')

from selenium import webdriver
from pyquery import PyQuery
import requests
import time
import datetime
import bs4
import re
# driver_path = r'D:\driver\chromedriver.exe'
# driver = webdriver.Chrome(executable_path=driver_path)
# url = "https://www.gcable.com.cn/about-us/purchase/?json=1"
# headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
#                 'Cookie': 'acw_tc=76b20f4516110367979744225e9ddf97264f252545a437e5d152f01c5d7a85; ECSID=3dc9c8cc66d131b20bccb64c6d6a9230'}
# res=requests.post(url, verify=False,headers=headers)
# soup=res.json()
# today = time.strftime('%Y-%m-%d', time.localtime())#获取当前时间
# d1 = datetime.datetime.strptime(today, '%Y-%m-%d')
# for item in soup:
#     d2 = datetime.datetime.strptime(item['post-time'], '%Y-%m-%d')
#     d=d1-d2
#     if (d.days < 11) & (item['type']=='采购公告'):
#         print(item)
# bmjzsj=['投标文件提交截止时间：','投标截止时间为','递交的截止时间','投标截止时间']
# for item in bmjzsj:
#     g = soup.text.find(item)
#     if g!=-1:
#         break
# print(g)
# g = soup.text.find('2', g, g + 25)
# h = soup.text.find('分', g)
# endTime = soup.text[g:h]
# endTime = endTime.replace(' ', '')
# endTime = endTime.replace('\n', '')
# print(soup.text)
# print(endTime)
# print(g)
# print(h)
